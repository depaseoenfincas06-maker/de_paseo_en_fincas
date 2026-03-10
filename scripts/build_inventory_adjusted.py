from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE_PATH = Path("/Users/juandavidvizcaya/Downloads/JP BASE DE DATOS FINCAS ACTUALIZADAS.xlsx")
OUTPUT_DIR = Path("/Users/juandavidvizcaya/Desktop/mscpn8n paseoenfinacas/output/spreadsheet")
OUTPUT_XLSX = OUTPUT_DIR / "JP_BASE_DE_DATOS_FINCAS_AJUSTADA_N8N.xlsx"
OUTPUT_CSV = OUTPUT_DIR / "fincas_inventory_ajustada_real.csv"


REQUIRED_COLUMNS = [
    "finca_id",
    "nombre",
    "zona",
    "municipio",
    "activa",
    "prioridad",
    "capacidad_max",
    "min_noches",
    "precio_noche_base",
    "precio_fin_semana",
    "deposito_seguridad",
    "precio_persona_extra",
    "pet_friendly",
    "amenidades_csv",
    "tipo_evento_csv",
    "descripcion_corta",
    "foto_url",
    "owner_nombre",
    "owner_contacto",
    "descuento_max_pct",
]

EXTRA_COLUMNS = [
    "source_row",
    "codigo_original",
    "zona_original",
    "capacidad_minima_tarifa",
    "precio_festivo",
    "precio_semana_santa_receso",
    "precio_temporada_alta",
    "habitaciones",
    "empleadas",
    "especificacion_habitaciones",
    "observaciones_originales",
    "caracteristicas_originales",
    "administrador_nombre",
    "administrador_contacto",
    "pricing_model",
    "review_status",
    "review_notes",
]

OUTPUT_COLUMNS = REQUIRED_COLUMNS + EXTRA_COLUMNS

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_FILLS = {
    "READY_FOR_OFFERING": PatternFill(fill_type="solid", fgColor="E2F0D9"),
    "REVIEW_PRICE_MODEL": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "INCOMPLETE": PatternFill(fill_type="solid", fgColor="FCE4D6"),
}


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"\n{2,}", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def title_case_location(text: str) -> str:
    if not text:
        return ""
    words = clean_text(text).split()
    keep_lower = {"de", "del", "la", "las", "los", "y"}
    out: list[str] = []
    for idx, word in enumerate(words):
        low = word.lower()
        if idx > 0 and low in keep_lower:
            out.append(low)
        else:
            out.append(low.capitalize())
    return " ".join(out)


def normalize_zone_and_municipio(code: str) -> tuple[str, str]:
    raw = re.split(r"\s*#", clean_text(code))[0].strip()
    raw_title = title_case_location(raw)

    zone_map = {
        "Pereira": ("Eje cafetero", "Pereira"),
        "Quindio": ("Eje cafetero", "Quindio"),
        "Eje Cafetero": ("Eje cafetero", "Eje cafetero"),
        "Santafe": ("Antioquia", "Santa Fe de Antioquia"),
        "Sopetran": ("Antioquia", "Sopetran"),
        "San Jeronimo": ("Antioquia", "San Jeronimo"),
        "Guatape": ("Antioquia", "Guatape"),
        "Antioquia": ("Antioquia", "Antioquia"),
        "Carmen De Apicala": ("Carmen de Apicala", "Carmen de Apicala"),
        "Girardot M2": ("Girardot", "Girardot"),
        "Girardot Q12": ("Girardot", "Girardot"),
    }
    if raw_title in zone_map:
        return zone_map[raw_title]
    return raw_title, raw_title


def parse_int(value: object) -> int | None:
    if value is None:
        return None
    text = clean_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+", text.replace(".", "").replace(",", ""))
    if not match:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def parse_money(value: object) -> dict[str, object]:
    text = clean_text(value).upper()
    if not text:
        return {"value": None, "per_person": False, "raw": ""}

    per_person = any(token in text for token in ["X PX", "POR PX", "POR PERSONA", "P/PX", "PAX"])

    normalized = text.replace("$", "").replace("COP", "").replace(" ", "")
    if "MIL" in normalized:
        match = re.search(r"(\d+(?:[.,]\d+)?)", normalized)
        if match:
            base = float(match.group(1).replace(".", "").replace(",", "."))
            return {"value": int(base * 1000), "per_person": per_person, "raw": text}

    digits = re.sub(r"[^\d]", "", normalized)
    if digits:
        return {"value": int(digits), "per_person": per_person, "raw": text}

    return {"value": None, "per_person": per_person, "raw": text}


def parse_contact(value: object) -> tuple[str, str]:
    text = clean_text(value)
    if not text:
        return "", ""
    phone_match = re.search(r"(?:\+?57[\s-]*)?(3\d{2})[\s-]*(\d{3})[\s-]*(\d{4})", text)
    phone = ""
    if phone_match:
        phone = "+57" + "".join(phone_match.groups())
        text = re.sub(r"(?:\+?57[\s-]*)?(3\d{2})[\s-]*(\d{3})[\s-]*(\d{4})", "", text).strip(" -/,;")
    return clean_text(text), phone


def infer_pet_friendly(text: str) -> bool:
    sample = text.lower()
    if "no mascota" in sample or "no pets" in sample:
        return False
    return any(token in sample for token in ["mascota", "pet friendly", "pets"])


def infer_amenidades(text: str, empleadas: str) -> list[str]:
    sample = text.lower()
    mapping = [
        ("piscina", ["piscina"]),
        ("jacuzzi", ["jacuzzi"]),
        ("bbq", ["bbq"]),
        ("wifi", ["wifi", "wi-fi"]),
        ("aire_acondicionado", ["aire acondicionado", "a.c", "ac."]),
        ("kiosco", ["kiosco"]),
        ("cancha", ["cancha"]),
        ("parqueadero", ["parqueadero", "garaje"]),
        ("turco", ["turco"]),
        ("sauna", ["sauna"]),
        ("billar", ["billar", "pool"]),
        ("ping_pong", ["ping pong"]),
        ("parque_infantil", ["parque infantil"]),
        ("lago", ["lago"]),
        ("rio", ["rio", "quebrada"]),
        ("discoteca", ["discoteca"]),
        ("bar", ["bar"]),
        ("zonas_verdes", ["zonas verdes"]),
        ("mirador", ["mirador"]),
        ("tenis", ["tennis", "tenis"]),
        ("golf", ["golf"]),
    ]
    found: list[str] = []
    for key, needles in mapping:
        if any(needle in sample for needle in needles):
            found.append(key)
    if clean_text(empleadas).upper().startswith("SI"):
        found.append("empleada")
    return sorted(dict.fromkeys(found))


def infer_tipo_evento(text: str) -> list[str]:
    sample = text.lower()
    found: list[str] = []
    if any(token in sample for token in ["familia", "familiar", "plan familiar", "solo familia"]):
        found.append("familiar")
    if any(token in sample for token in ["fiesta", "rumba"]):
        found.append("fiesta")
    if "evento empresarial" in sample or "empresarial" in sample:
        found.append("evento_empresarial")
    if any(token in sample for token in ["matrimonio", "matrimonios", "boda"]):
        found.append("matrimonio")
    if not found or any(token in sample for token in ["descanso", "solo descanso"]):
        found.append("descanso")
    return sorted(dict.fromkeys(found))


def shorten_description(text: str, limit: int = 220) -> str:
    clean = clean_text(text)
    if len(clean) <= limit:
        return clean
    shortened = clean[:limit].rsplit(" ", 1)[0].strip()
    return shortened + "..."


def build_review_status(name: str, zona: str, capacidad: int | None, price_info: dict[str, object]) -> tuple[str, str, bool]:
    notes: list[str] = []
    if not name:
        notes.append("falta nombre")
    if not zona:
        notes.append("falta zona")
    if capacidad is None:
        notes.append("falta capacidad_max")

    if price_info["per_person"]:
        notes.append("tarifa origen en formato por persona")
    elif price_info["value"] is None:
        notes.append("sin tarifa FDS normal usable")
    else:
        notes.append("precio_noche_base vacio; origen trae tarifa FDS/paquete")

    if not name or not zona or capacidad is None:
        return "INCOMPLETE", "; ".join(notes), False
    if price_info["per_person"] or price_info["value"] is None:
        return "REVIEW_PRICE_MODEL", "; ".join(notes), True
    return "READY_FOR_OFFERING", "; ".join(notes), True


def copy_sheet_values(src_ws, dest_ws) -> None:
    for row in src_ws.iter_rows():
        for cell in row:
            dest_ws[cell.coordinate] = cell.value
    for key, dim in src_ws.column_dimensions.items():
        if dim.width:
            dest_ws.column_dimensions[key].width = dim.width


def autosize(ws) -> None:
    for idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    source_wb = load_workbook(SOURCE_PATH, data_only=True)
    source_ws = source_wb[source_wb.sheetnames[0]]

    rows_out: list[dict[str, object]] = []
    zone_counter: Counter[str] = Counter()
    status_counter: Counter[str] = Counter()

    for row_idx in range(4, source_ws.max_row + 1):
        source_values = [source_ws.cell(row_idx, col_idx).value for col_idx in range(1, 17)]
        if not any(value not in (None, "") for value in source_values[:16]):
            continue

        codigo_original = clean_text(source_values[0])
        nombre = clean_text(source_values[1])
        capacidad_max = parse_int(source_values[2])
        capacidad_minima_tarifa = parse_int(source_values[3])
        tarifa_fds = parse_money(source_values[4])
        tarifa_festivo = parse_money(source_values[5])
        tarifa_semana_santa = parse_money(source_values[6])
        tarifa_temp_alta = parse_money(source_values[7])
        observaciones = clean_text(source_values[8])
        caracteristicas = clean_text(source_values[9])
        empleadas = clean_text(source_values[10])
        habitaciones = parse_int(source_values[11])
        especificacion_habitaciones = clean_text(source_values[12])
        foto_url = clean_text(source_values[13])
        administrador_nombre, administrador_contacto = parse_contact(source_values[14])
        owner_nombre, owner_contacto = parse_contact(source_values[15])
        zona, municipio = normalize_zone_and_municipio(codigo_original)
        zona_original = re.split(r"\s*#", codigo_original)[0].strip() if codigo_original else ""
        merged_text = f"{observaciones}\n{caracteristicas}"
        amenidades = infer_amenidades(merged_text, empleadas)
        tipo_evento = infer_tipo_evento(merged_text)
        pet_friendly = infer_pet_friendly(merged_text)
        review_status, review_notes, activa = build_review_status(nombre, zona, capacidad_max, tarifa_fds)

        precio_fin_semana = ""
        precio_persona_extra = ""
        pricing_model = ""
        if tarifa_fds["per_person"]:
            precio_persona_extra = tarifa_fds["value"] or ""
            pricing_model = "por_persona"
        elif tarifa_fds["value"] is not None:
            precio_fin_semana = tarifa_fds["value"]
            pricing_model = "paquete_fin_de_semana"
        else:
            pricing_model = "sin_tarifa_clara"

        row = {
            "finca_id": codigo_original.replace(" ", "_"),
            "nombre": nombre,
            "zona": zona,
            "municipio": municipio,
            "activa": "true" if activa else "false",
            "prioridad": len(rows_out) + 1,
            "capacidad_max": capacidad_max or "",
            "min_noches": 2 if activa else "",
            "precio_noche_base": "",
            "precio_fin_semana": precio_fin_semana,
            "deposito_seguridad": 0 if activa else "",
            "precio_persona_extra": precio_persona_extra,
            "pet_friendly": "true" if pet_friendly else "false",
            "amenidades_csv": ",".join(amenidades),
            "tipo_evento_csv": ",".join(tipo_evento),
            "descripcion_corta": shorten_description(caracteristicas or observaciones),
            "foto_url": foto_url,
            "owner_nombre": owner_nombre,
            "owner_contacto": owner_contacto,
            "descuento_max_pct": 0 if activa else "",
            "source_row": row_idx,
            "codigo_original": codigo_original,
            "zona_original": zona_original,
            "capacidad_minima_tarifa": capacidad_minima_tarifa or "",
            "precio_festivo": tarifa_festivo["value"] or "",
            "precio_semana_santa_receso": tarifa_semana_santa["value"] or "",
            "precio_temporada_alta": tarifa_temp_alta["value"] or "",
            "habitaciones": habitaciones or "",
            "empleadas": empleadas,
            "especificacion_habitaciones": especificacion_habitaciones,
            "observaciones_originales": observaciones,
            "caracteristicas_originales": caracteristicas,
            "administrador_nombre": administrador_nombre,
            "administrador_contacto": administrador_contacto,
            "pricing_model": pricing_model,
            "review_status": review_status,
            "review_notes": review_notes,
        }
        rows_out.append(row)
        zone_counter[zona] += 1
        status_counter[review_status] += 1

    csv_ready_rows = [{column: row.get(column, "") for column in OUTPUT_COLUMNS} for row in rows_out]

    with OUTPUT_CSV.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(csv_ready_rows)

    wb = Workbook()
    readme_ws = wb.active
    readme_ws.title = "README"
    normalized_ws = wb.create_sheet("fincas_n8n")
    source_copy_ws = wb.create_sheet("source_snapshot")

    readme_lines = [
        ["Documento", "JP BASE DE DATOS FINCAS AJUSTADA N8N"],
        ["Fuente", str(SOURCE_PATH)],
        ["Hoja fuente", source_ws.title],
        ["Filas detectadas", len(rows_out)],
        ["READY_FOR_OFFERING", status_counter["READY_FOR_OFFERING"]],
        ["REVIEW_PRICE_MODEL", status_counter["REVIEW_PRICE_MODEL"]],
        ["INCOMPLETE", status_counter["INCOMPLETE"]],
        ["Regla 1", "Se preservan las fincas reales y se normalizan al esquema de n8n."],
        ["Regla 2", "precio_noche_base se deja vacio porque la fuente actual maneja principalmente tarifas FDS/paquete, no tarifa por noche."],
        ["Regla 3", "precio_fin_semana toma TARIFA FDS NORMAL cuando es numerica y total."],
        ["Regla 4", "precio_persona_extra se usa cuando la tarifa origen viene por persona."],
        ["Regla 5", "min_noches se dejo en 2 como default operativo para esta fase porque la fuente no trae ese dato."],
        ["Regla 6", "capacidad_minima_tarifa conserva CANTIDAD MINIMA del archivo original para revision comercial."],
        ["Regla 7", "zona se normalizo a zonas comerciales amplias. municipio conserva la ubicacion original util."],
        ["Regla 8", "activa=false solo en filas incompletas. REVIEW_PRICE_MODEL sigue activa para offering pero requiere revision antes de negociar automaticamente."],
        ["Zonas detectadas", ", ".join(f"{zone} ({count})" for zone, count in zone_counter.most_common())],
    ]
    for row_idx, row in enumerate(readme_lines, start=1):
        readme_ws.cell(row_idx, 1).value = row[0]
        readme_ws.cell(row_idx, 2).value = row[1]
    readme_ws.column_dimensions["A"].width = 22
    readme_ws.column_dimensions["B"].width = 140
    readme_ws.freeze_panes = "A2"

    for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
        cell = normalized_ws.cell(1, col_idx)
        cell.value = column_name
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    status_col_idx = OUTPUT_COLUMNS.index("review_status") + 1
    for row_idx, row in enumerate(csv_ready_rows, start=2):
        for col_idx, column_name in enumerate(OUTPUT_COLUMNS, start=1):
            normalized_ws.cell(row_idx, col_idx).value = row[column_name]
        status_cell = normalized_ws.cell(row_idx, status_col_idx)
        status_fill = STATUS_FILLS.get(str(status_cell.value))
        if status_fill:
            status_cell.fill = status_fill

    normalized_ws.freeze_panes = "A2"
    normalized_ws.auto_filter.ref = f"A1:{get_column_letter(normalized_ws.max_column)}{normalized_ws.max_row}"
    autosize(normalized_ws)

    copy_sheet_values(source_ws, source_copy_ws)
    source_copy_ws.title = "source_snapshot"
    source_copy_ws.freeze_panes = "A2"

    wb.save(OUTPUT_XLSX)

    print(f"Wrote {OUTPUT_XLSX}")
    print(f"Wrote {OUTPUT_CSV}")
    print(f"rows={len(rows_out)}")
    print(f"status_counts={dict(status_counter)}")


if __name__ == "__main__":
    main()
